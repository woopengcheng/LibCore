#-*- coding: utf-8 -*-
#coding:utf-8
import sys , getopt
import os , re , string
import time , datetime
import binascii 
import collections  
import xml.dom.minidom
import time,datetime
import socket
import csv

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook

from xml.dom import minidom , Node 

from xml.etree.ElementTree import ElementTree
from xml.etree.ElementTree import Element
from xml.etree.ElementTree import SubElement as SE
#sys.reload()
#sys.setdefaultencoding("utf-8") 
#sys.setdefaultencoding("cp936") 

g_xlsImportPath = ""
g_xlsExportCSVPath = ""
g_xlsExportCPPPath = ""

g_xlsRecords = {}
g_configPrefix = "S"
g_loadConfigSuffix = "Load"
g_xlsNamespace = "Config"
g_int32Array = 1
g_int64Array = 2
g_boolArray = 3
g_doubleArray = 4
g_stringArray = 5
	
oneTab = "\t"
twoTab = oneTab + "\t"
threeTab = twoTab + "\t"
fourTab = threeTab + "\t"
fiveTab = fourTab + "\t"
sixTab = fiveTab + "\t"

def start(): 
	LogOutInfo("start generate csv.\n")   
	CreateExportPathFiles()
	DeleteExportPathFiles()
	GenerateCSVFromXLS()
	LogOutInfo("generate CSV finished.\n") 
			
	LogOutInfo("start generate CPP.\n")   	
	GenerateCPP()
	LogOutInfo("generate CPP finished.\n") 

def GenerateCSVFromXLS():
	root = g_xlsImportPath
	
	files = Search(root ,'.xlsx')
	for result in files:
		LogOutInfo("filename:" , result);
		Xlsx2CSV(result)
		
def GenerateCPP(): 
	GenerateConfigManagerHeader()
	for sheet , item in g_xlsRecords.items():
		GenerateConfigLoadHeader(sheet , item[2] , item[1] , item[0])
		GenerateConfigLoadCpp(sheet , item[2] , item[1] , item[0])
		GenerateConfigHeader(sheet , item[2] , item[1] , item[0])
		GenerateConfigCpp(sheet , item[2] , item[1] , item[0])
		for row , item2 in g_xlsRecords[sheet].items():
			#LogOutInfo("row" , row , "size:" , len(item2) , "size2" , len(item2[2]), "size1" , len(item2[1]), "size0" , len(item2[0]))
			#LogOutInfo("row" , row)
			pass
			#GenerateConfigHeader(sheet , item2[2] , item2[1] , item2[0])
	GenerateConfigManagerCPP()

def GenerateConfigLoadHeader(filename , types , datas , comments):
	filename = filename + g_loadConfigSuffix
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".h"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	#以下是生成数据定义
	WriteFileDescription(fileWrite , filename + ".h" , "csv配置文件")
	fileWrite.write("#ifndef __" + filename + "_define_h__\n")
	fileWrite.write("#define __" + filename + "_define_h__\n") 
	fileWrite.write("#include \"CUtil/inc/Common.h \"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "struct " + g_configPrefix + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	 
	for index , item in enumerate(types):
		fileWrite.write(twoTab + GetType(item) + GetTypeTab(item) + datas[index] + ";"  + oneTab)
		fileWrite.write("//" + comments[index] + "\n") 
	fileWrite.write(oneTab + "};\n\n\n") 
			
	#以下是生成导入数据接口
	fileWrite.write(oneTab + "class " + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "typedef std::vector<" + g_configPrefix + filename + "> CollectionConfigsT;\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "bool LoadFrom(const char* filename);\n")
	fileWrite.write(twoTab + "bool LoadFrom(const std::string& filename)")
	fileWrite.write("{ ")
	fileWrite.write("return LoadFrom(filename.c_str()); ")
	fileWrite.write("}\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + g_configPrefix + filename + " & Get(size_t row);\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "inline size_t Count()")
	fileWrite.write("{ ")
	fileWrite.write("return m_vtConfigs.size(); ")
	fileWrite.write("}\n\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "CollectionConfigsT m_vtConfigs;\n")	
	fileWrite.write(oneTab + "};\n") 


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + filename + "_define_h__\n") 

	fileWrite.close()	  
	
def GenerateConfigLoadCpp(filename , types , datas , comments):
	filename = filename + g_loadConfigSuffix
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	WriteFileDescription(fileWrite , filename + ".cpp" , "csv读取文件实现")
	fileWrite.write("#include \"" + filename + ".h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CUtil.h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CSVReader.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "bool " + filename + "::LoadFrom(const char* filename)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + "CUtil::CSVReader csv;\n") 
	fileWrite.write(twoTab + "if(csv.Load(filename) != 0)\n") 
	fileWrite.write(threeTab + "return false;\n\n") 

	for index , item in enumerate(types):
		fileWrite.write(twoTab + "size_t index_" + datas[index] + " = csv.GetIndex(\"" + datas[index] + "\", 1);\n") 
		fileWrite.write(twoTab + "MsgAssert_Re0(index_" + datas[index] + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 

	fileWrite.write(twoTab + "for (size_t row = 3; row < csv.Count(); ++row)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_configPrefix + filename + " conf;\n\n") 
	for index , item in enumerate(types):
		Str = GetTypeFunc(item)
		if type(Str) != str :
			if Str == g_int32Array:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((INT32)CUtil::atoi(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_int64Array:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((INT64)CUtil::atoi(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_boolArray:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((bool)CUtil::atoi(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_doubleArray:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((float)CUtil::atof(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_stringArray:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, conf." + datas[index] + ", \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(threeTab + "}\n\n") 
		else:
			fileWrite.write(threeTab + "conf." + datas[index] + " = csv." + Str + "(row , index_" + datas[index] + ");\n") 
			

	fileWrite.write(threeTab + "m_vtConfigs.push_back(conf);\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
	fileWrite.write(oneTab + g_configPrefix + filename + " & " + filename + "::Get(size_t row)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "return m_vtConfigs.at(row);\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write("}\n\n") 
	
	fileWrite.close()	

def GenerateConfigHeader(filename , types , datas , comments):
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".h"
	if os.path.exists(outputPath): 
		os.remove(outputPath)
#		return 

	fileWrite = open(outputPath , "a")
	
	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename

	WriteFileDescription(fileWrite , filename + ".h" , "csv读取文件")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_" + filename + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_" + filename +  "_define_h__\n") 
	fileWrite.write("#include \"" + loadConfig + ".h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 

	fileWrite.write(oneTab + "struct " + dataConfig + "\n") 
	fileWrite.write(oneTab + "{\n") 
	 
	for index , item in enumerate(types):
		fileWrite.write(twoTab + GetType(item) + GetTypeTab(item) + datas[index] + ";"  + oneTab)
		fileWrite.write("//" + comments[index] + "\n") 
	fileWrite.write(oneTab + "};\n\n\n") 
			
	fileWrite.write(oneTab + "class " + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "typedef std_unordered_map<" + GetType(types[0]) + " , " + dataConfig + "> MapConfigsT;\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "bool LoadFrom(const std::string& filepath);\n")
	fileWrite.write(twoTab + dataConfig + " * Get" + filename + "(INT32 nIndex);\n\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "MapConfigsT m_mapConfigs;\n\n")
	
	fileWrite.write(oneTab + "};\n") 			
	fileWrite.write(oneTab + "extern " + filename + " * g_p" + filename + ";\n") 


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_" + filename + "_define_h__\n") 
	
	fileWrite.close()	 	


def GenerateConfigCpp(filename , types , datas , comments):
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename
	fileWrite = open(outputPath , "a")
	
	WriteFileDescription(fileWrite , filename + ".cpp" , "csv读取数据文件实现")
	fileWrite.write("#include \"" + filename + ".h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "bool " + filename + "::LoadFrom(const std::string & filepath)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + g_xlsNamespace + "::" + loadConfig + " loadConfig;\n") 
	fileWrite.write(twoTab + "MsgAssert_Re0(loadConfig.LoadFrom(filepath) , \"Error " + filename + "LoadFrom \" << filepath);\n\n") 
	
	fileWrite.write(twoTab + "for(size_t i = 0; i < loadConfig.Count(); ++i)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + g_configPrefix + loadConfig + "& config = loadConfig.Get(i);\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + dataConfig + " data = {0};\n") 
	
	for index , item in enumerate(types):
		fileWrite.write(threeTab + "data." + datas[index] + " = config." + datas[index] + ";\n") 

	fileWrite.write(threeTab + "m_mapConfigs.insert(std::make_pair(data." + datas[0] + ",data));\n") 
	fileWrite.write(twoTab + "}\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
	fileWrite.write(oneTab + dataConfig + " * " + filename + "::Get" + filename + "(INT32 nIndex)\n")
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MapConfigsT::iterator iter = m_mapConfigs.find(nIndex);\n") 
	fileWrite.write(twoTab + "if(iter == m_mapConfigs.end())\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "gWarniStream( \"" + filename + "::Get" + filename + " NotFound \" << nIndex);\n") 
	fileWrite.write(threeTab + "return NULL;\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return &iter->second;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + filename + " * g_p" + filename + " = NULL;\n") 
	fileWrite.write("}\n\n") 
	
	fileWrite.close()

def GenerateConfigManagerHeader():
	outputPath = g_xlsExportCPPPath  + os.sep + "ConfigManager.h"
	if os.path.exists(outputPath): 
		os.remove(outputPath) 

	fileWrite = open(outputPath , "a") 

	WriteFileDescription(fileWrite , "ConfigManager.h" , "ConfigManager文件声明")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_ConfigManager" + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_ConfigManager" +  "_define_h__\n") 
	fileWrite.write("#include \"CUtil/inc/Common.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 
			
	fileWrite.write(oneTab + "class ConfigManager\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "ConfigManager();\n")
	fileWrite.write(twoTab + "~ConfigManager();\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "static ConfigManager & GetInstance();\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "INT32		Init(std::string  strCsvPath);\n")
	fileWrite.write(twoTab + "INT32		Cleanup();\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "\n\n")
	fileWrite.write(oneTab + "};\n") 		


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_ConfigManager_define_h__\n") 
	
	fileWrite.close()	 	


def GenerateConfigManagerCPP():
	outputPath = g_xlsExportCPPPath  + os.sep + "ConfigManager.cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	WriteFileDescription(fileWrite , "ConfigManager.cpp" , "ConfigManager数据管理文件实现")
	fileWrite.write("#include \"" + "ConfigManager.h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n") 
	for sheet , item in g_xlsRecords.items():
		fileWrite.write("#include \"" + sheet + ".h\"\n") 
	fileWrite.write(oneTab + "\n\n") 

	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "ConfigManager::ConfigManager()\n") 
	fileWrite.write(oneTab + "{\n") 
	for sheet , item in g_xlsRecords.items():
		fileWrite.write(twoTab + "g_p" + sheet + " = new " + g_xlsNamespace + "::" + sheet + ";\n") 
	fileWrite.write(oneTab + "}\n\n") 
	
	fileWrite.write(oneTab + "ConfigManager::~ConfigManager()\n") 
	fileWrite.write(oneTab + "{\n") 
	for sheet , item in g_xlsRecords.items():
		fileWrite.write(twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  "g_p" + sheet + ");\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "ConfigManager & ConfigManager::GetInstance()\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "static ConfigManager instance;\n") 
	fileWrite.write(twoTab + "return instance;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "INT32 ConfigManager::Init(std::string  strCsvPath)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MsgAssert_ReF1(strCsvPath.length(), \"ConfigManager::Init error.\");\n\n") 
	fileWrite.write(twoTab + "if (strCsvPath[strCsvPath.length() - 1] != '/')\n") 
	fileWrite.write(twoTab + "{\n") 
	for sheet , item in g_xlsRecords.items():
		fileWrite.write(threeTab + "MsgAssert_ReF1(" + g_xlsNamespace + "::" +  "g_p" + sheet + " , \"ConfigManager not Init\")\n") 
		fileWrite.write(threeTab + "" + g_xlsNamespace + "::" +  "g_p" + sheet + "->LoadFrom(strCsvPath + \"" + sheet + ".tabcsv\");\n\n") 
	fileWrite.write(twoTab + "}\n") 
	fileWrite.write(twoTab + "return 0;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "INT32 ConfigManager::Cleanup()\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "return -1;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write("}\n\n") 
	
	fileWrite.close()

################################流程相关函数处理#####################################
def RemoveNewLine(str):
	pp = ""
	for tt in str.splitlines():
		tt = tt.rstrip()
		pp = pp+tt
	return pp

def RemovListNewLine(List): 
	for index , item2 in enumerate(List):
		List[index] = RemoveNewLine(item2)

def GetType(item):
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return "INT32"
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return "INT64"
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return "double"
	elif item.lower() == "bool".lower():
		return "bool"
	elif item.lower() == "string".lower():
		return "std::string"
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return "std::vector<INT32>" 
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return "std::vector<INT64>"
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return "std::vector<double>"
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return "std::vector<std::string>"
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return "std::vector<bool>"
	else:
		LogOutError("GetType error." , item)

	return "None"

def GetTypeFunc(item):
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return "GetInt32"
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return "GetInt64"
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return "GetDouble"
	elif item.lower() == "bool".lower():
		return "GetBool"
	elif item.lower() == "string".lower():
		return "GetString"
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return g_int32Array
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return g_int64Array
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return g_doubleArray
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return g_stringArray
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return g_boolArray
	else:
		LogOutError("GetTypeFunc error." , item)

	return None

def GetTypeTab(item):
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return sixTab
	elif item.lower() == "bool".lower():
		return sixTab
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return sixTab
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return sixTab
	elif item.lower() == "string".lower():
		return fiveTab
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return threeTab
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return threeTab
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return threeTab
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return threeTab
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return oneTab
	else:
		LogOutError("GetTypeTab error." , item)

	return oneTab


################################流程无关函数处理#####################################
def Usage():
    print('GenerateCSV.py usage:')
    print('-h,--help: print help message.')
    print('-v, --version: print script version')
    print('-o, --output: input an csv output verb')
    print('-c, --output: input an cpp output verb')
    print('--foo: Test option ')
    print('--fre: another test option')

def Version():
	print('GenerateCSV.py 1.0.0.0.1')

def LogOutDebug(*string):
	longStr = "debug: "
	for item in range(len(string)):  
		longStr += str(string[item])

	print(longStr)
	pass

def LogOutInfo(*string):
	longStr = "info: "
	for item in range(len(string)):  
		longStr += str(string[item])
	
	print(longStr)
	
def LogOutError(*string):
	longStr = "error: "
	for item in range(len(string)):  
		longStr += str(string[item])
	
	print(longStr)
	sys.exit()
	
def WriteFileDescription(fileWrite , sfile , desc):
	fileWrite.write("/************************************" + "\n")
	fileWrite.write("FileName	:	" + sfile + "\n")
	fileWrite.write("Author		:	generate by tools" + "\n")
	fileWrite.write("HostName	:	" + socket.gethostname() + "\n")
	fileWrite.write("IP			:	" + socket.gethostbyname(socket.gethostname()) + "\n")
	fileWrite.write("Version		:	0.0.1" + "\n")
#	fileWrite.write("Date		:	" + time.strftime('%Y-%m-%d %H:%M:%S') + "\n")
	fileWrite.write("Description	:	" + desc + "\n")
	fileWrite.write("************************************/" + "\n")

def Search(base , suffix):
    pattern = suffix
    fileresult = []
    cur_list = os.listdir(base)
    for item in cur_list:
        full_path = os.path.join(base, item)
        if os.path.isdir(full_path):            
            fileresult += Search(full_path , suffix)
        if os.path:
            if full_path.endswith(pattern):
                fileresult.append(full_path)
    return fileresult
	
def DeleteExportPathFiles():
	files = Search(g_xlsExportCSVPath , '.tabcsv')
	for sfile in files:
		if os.path.exists(sfile): 
			os.remove(sfile)

def CreateExportPathFiles(): 
	if False == os.path.exists(g_xlsExportCSVPath):
		os.makedirs(g_xlsExportCSVPath)
		LogOutInfo("create dir: " , g_xlsExportCSVPath)
		
	if False == os.path.exists(g_xlsExportCPPPath):
		os.makedirs(g_xlsExportCPPPath)
		LogOutInfo("create dir: " , g_xlsExportCPPPath)

def Xlsx2CSV(filepath):
	dirout = g_xlsExportCSVPath
	dirout = dirout + os.sep  #新路径名称
	try:
		# 一个表中的所有sheet输出到一个csv文件中,所以要保证格式一致.文件名用xlsx文件名
		filename = os.path.basename(filepath) #获取文件名
		filename = os.path.splitext(filename.replace(' ', '_'))[0]
		csv_filename = '{xlsx}.tabcsv'.format(xlsx=filename)		
		dirfileout  = dirout + csv_filename
		#LogOutDebug("dirfileout" , dirfileout )
		csv_file = open(dirfileout , 'w' , newline='')
		
		#QUOTE_MINIMAL QUOTE_NONE所有的都要加引号.
		csv_file_writer = csv.writer(csv_file , delimiter='	', quotechar='"', quoting=csv.QUOTE_ALL)
		
		id_list = [] #用于重复的ID去除
		cur_sheet_index = 0
		xlsx_file_reader = load_workbook(filepath)
		g_xlsRecords[filename] = {}
		for sheet in xlsx_file_reader.get_sheet_names():			
			cur_rows_index = 0
			sheet_ranges = xlsx_file_reader[sheet]
			for row in sheet_ranges.rows:
				if cur_sheet_index >= 1 and cur_rows_index < 4:
					cur_rows_index = cur_rows_index + 1
					continue
				
				row_container = []
				cur_cell_index = 0
				for cell in row:
					if type(cell.value) != type(None):
						Str = cell.value
						if type(cell.value) != str:
							Str = str(cell.value)
						else:
							Str = Str.encode('gbk').decode('gbk')
						if len(Str) >= 1:
							cur_cell_index = cur_cell_index + 1
							if cur_cell_index == 1:		# 插入ID
								if Str in id_list:
									LogOutDebug("repeat id \'" , Str , "\' in \'" , filepath , " \' file")
									#LogOutError("repeat id \'" , Str , "\' in \'" , filepath , " \' file")
								id_list.append(Str)
							row_container.append(Str)
						#LogOutDebug("cell.value:" , type(cell.value) , Str)
					
				if len(row_container) >= 1:	
					RemovListNewLine(row_container)
					g_xlsRecords[filename] [cur_rows_index] = row_container
					csv_file_writer.writerow(row_container)
					cur_rows_index = cur_rows_index + 1
					#LogOutDebug("cell.row_container:" , row_container)
				
			cur_sheet_index = cur_sheet_index + 1		
		csv_file.close()
	except Exception as e:
		LogOutError(e)
			
################################main函数处理#####################################
def handleArgs(argv): 
	global g_xlsImportPath 
	global g_xlsExportCSVPath
	global g_xlsExportCPPPath
	
	try:
		opts, args = getopt.getopt(argv[1:], 'hvi:o:c:', ['import='])
	except: 
		Usage()
		sys.exit(2) 
	if len(argv) < 2: 
		Usage()
		return  
	for o, a in opts:
		if o in ('-h', '--help'):
			Usage()
			sys.exit(1)
		elif o in ('-v', '--version'):
			Version()
			sys.exit(0) 
		elif o in ('-i','--import',):
			g_xlsImportPath = a
		elif o in ('-o','--export',):
			g_xlsExportCSVPath = a
		elif o in ('-c','--cppexport',):
			g_xlsExportCPPPath = a
		elif o in ('--fre',):
			Fre=a
		else:
			print('unhandled option')
			sys.exit(3) 
			
def main(argv):
	global g_xlsImportPath 
	global g_xlsExportCSVPath
	global g_xlsExportCPPPath
#	handleArgs(argv)
	g_xlsImportPath = "./xls_config"
	g_xlsExportCSVPath = "../../../bin/vs14.0/x64/DLL_Debug_x64/csv_config"
	g_xlsExportCPPPath = "../../../vsproject/TestLibCore/CSVConfigs"
	LogOutInfo("generate csv from path:" + g_xlsImportPath + " csv will export to:" + g_xlsExportCSVPath + " cpp will export to:" + g_xlsExportCPPPath) 
	start()
	LogOutInfo("complete generate csv.") 
	
if __name__ == '__main__': 
	main(sys.argv)


